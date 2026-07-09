from io import BytesIO

import pytest
from mpt_extension_contrib.reports import ExcelReportBuilder
from openpyxl import load_workbook


def test_build_sheet_writes_headers_and_rows():
    builder = ExcelReportBuilder(["ID", "Name"], sheet_name="Pending Orders")

    result = builder.build_sheet([["1", "Alice"], ["2", "Bob"]])

    worksheet = load_workbook(BytesIO(result)).active
    assert worksheet.title == "Pending Orders"
    assert [cell.value for cell in worksheet[1]] == ["ID", "Name"]
    assert worksheet[1][0].font.bold is True
    assert [cell.value for cell in worksheet[2]] == ["1", "Alice"]
    assert worksheet.auto_filter.ref == "A1:B3"


def test_build_sheet_with_no_rows():
    builder = ExcelReportBuilder(["ID"])

    result = builder.build_sheet([])

    worksheet = load_workbook(BytesIO(result)).active
    assert worksheet.title == "Report"
    assert worksheet.max_row == 1


def test_build_multi_sheet():
    builder = ExcelReportBuilder(["ignored"])

    result = builder.build_multi_sheet([
        ("First", ["A"], [["1"]]),
        ("Second", ["B"], [["2"]]),
    ])

    workbook = load_workbook(BytesIO(result))
    assert workbook.sheetnames == ["First", "Second"]
    assert workbook["Second"][1][0].value == "B"


def test_build_multi_sheet_rejects_empty():
    builder = ExcelReportBuilder(["ID"])

    with pytest.raises(ValueError):
        builder.build_multi_sheet([])


def test_save_writes_bytes_to_path(tmp_path):
    builder = ExcelReportBuilder(["ID"])
    target = tmp_path / "report.xlsx"

    result = builder.save(str(target), b"payload")

    assert result is None
    assert target.read_bytes() == b"payload"
