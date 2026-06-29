from dataclasses import dataclass
from io import BytesIO
from typing import ClassVar

import pytest
from freezegun import freeze_time
from mpt_api_client.rql import RQLQuery
from mpt_extension_contrib.reports import (
    EXCEL_MIME_TYPE,
    PENDING_ORDER_STATUSES,
    pending_orders,
    publish_xlsx_to_confluence,
    publish_xlsx_to_confluence_async,
)
from openpyxl import load_workbook


class FakeResource:
    def __init__(self, payload):
        self._payload = payload

    def to_dict(self):
        return self._payload


class FakeOrdersService:
    def __init__(self, payloads):
        self._payloads = payloads
        self.applied_query = None
        self.applied_select = None

    def filter(self, rql):
        self.applied_query = rql
        return self

    def select(self, *fields):
        self.applied_select = fields
        return self

    def iterate(self, batch_size=100):
        return iter([FakeResource(payload) for payload in self._payloads])


class FakeCommerce:
    def __init__(self, orders):
        self.orders = orders


class FakeClient:
    def __init__(self, orders):
        self.commerce = FakeCommerce(orders)


class FakeAsyncOrdersService:
    def __init__(self, payloads):
        self._payloads = payloads

    def filter(self, rql):
        return self

    def select(self, *fields):
        return self

    async def iterate(self, batch_size=100):
        for payload in self._payloads:
            yield FakeResource(payload)


@dataclass(frozen=True)
class FakeSettings:
    confluence_base_url: str = "https://example.atlassian.net/wiki"
    confluence_user: str = "user"
    confluence_token: str = "token"
    pending_orders_information_report_page_id: str = "page-1"


class FakeConfluence:
    result: ClassVar[bool] = True
    last: ClassVar["FakeConfluence | None"] = None

    def __init__(self, *, base_url, user, token):
        self.base_url = base_url
        self.user = user
        self.token = token
        self.call = None
        FakeConfluence.last = self

    def attach_content(self, page_id, filename, file_content, *, content_type, comment=None):
        self.call = {
            "page_id": page_id,
            "filename": filename,
            "file_content": file_content,
            "content_type": content_type,
            "comment": comment,
        }
        return FakeConfluence.result


@pytest.fixture
def confluence(monkeypatch):
    FakeConfluence.result = True
    FakeConfluence.last = None
    monkeypatch.setattr(pending_orders, "ConfluenceClient", FakeConfluence)
    return FakeConfluence


def _to_row(order):
    return [order["id"], order["status"]]


@freeze_time("2026-06-29")
def test_attaches_dated_workbook_and_comment(confluence):
    service = FakeOrdersService([
        {"id": "ORD-1", "status": "Querying"},
        {"id": "ORD-2", "status": "Processing"},
    ])

    result = publish_xlsx_to_confluence(
        FakeClient(service), FakeSettings(), row_mapper=_to_row, headers=["ID", "Status"]
    )

    assert result is True
    assert confluence.last.call["filename"] == "orders-29-06-2026.xlsx"
    assert confluence.last.call["content_type"] == EXCEL_MIME_TYPE
    assert confluence.last.call["comment"] == "Total orders 2"


def test_builds_confluence_and_page_from_settings(confluence):
    service = FakeOrdersService([{"id": "ORD-1", "status": "Querying"}])

    result = publish_xlsx_to_confluence(
        FakeClient(service), FakeSettings(), row_mapper=_to_row, headers=["ID", "Status"]
    )

    assert result is True
    assert confluence.last.base_url == "https://example.atlassian.net/wiki"
    assert confluence.last.user == "user"
    assert confluence.last.token == "token"
    assert confluence.last.call["page_id"] == "page-1"


def test_filters_pending_orders_and_builds_rows(confluence):
    service = FakeOrdersService([
        {"id": "ORD-1", "status": "Querying"},
        {"id": "ORD-2", "status": "Processing"},
    ])

    result = publish_xlsx_to_confluence(
        FakeClient(service), FakeSettings(), row_mapper=_to_row, headers=["ID", "Status"]
    )

    assert result is True
    assert service.applied_query == RQLQuery(status__in=list(PENDING_ORDER_STATUSES))
    assert service.applied_select == ("audit", "parameters")
    worksheet = load_workbook(BytesIO(confluence.last.call["file_content"])).active
    assert worksheet.title == "Pending Orders"
    assert [cell.value for cell in worksheet[2]] == ["ORD-1", "Querying"]


def test_infers_headers_and_rows(confluence):
    service = FakeOrdersService([{"id": "ORD-1", "status": "Querying"}])

    result = publish_xlsx_to_confluence(FakeClient(service), FakeSettings())

    assert result is True
    worksheet = load_workbook(BytesIO(confluence.last.call["file_content"])).active
    assert [cell.value for cell in worksheet[1]] == ["id", "status"]
    assert [cell.value for cell in worksheet[2]] == ["ORD-1", "Querying"]


def test_returns_false_when_upload_fails(confluence):
    confluence.result = False
    service = FakeOrdersService([{"id": "ORD-1", "status": "Querying"}])

    result = publish_xlsx_to_confluence(FakeClient(service), FakeSettings())

    assert result is False


def test_rejects_row_mapper_without_headers(confluence):
    client = FakeClient(FakeOrdersService([]))

    with pytest.raises(ValueError):
        publish_xlsx_to_confluence(client, FakeSettings(), row_mapper=_to_row)


def test_rejects_headers_without_row_mapper(confluence):
    client = FakeClient(FakeOrdersService([]))

    with pytest.raises(ValueError):
        publish_xlsx_to_confluence(client, FakeSettings(), headers=["ID"])


def test_select_is_overridable(confluence):
    service = FakeOrdersService([{"id": "X", "status": "Y"}])

    result = publish_xlsx_to_confluence(
        FakeClient(service), FakeSettings(), row_mapper=_to_row, headers=["ID"], select=["audit"]
    )

    assert result is True
    assert service.applied_select == ("audit",)


@freeze_time("2026-06-29")
async def test_async_attaches_report(confluence):
    client = FakeClient(
        FakeAsyncOrdersService([
            {"id": "ORD-1", "status": "Querying"},
            {"id": "ORD-2", "status": "Processing"},
        ])
    )

    result = await publish_xlsx_to_confluence_async(
        client, FakeSettings(), row_mapper=_to_row, headers=["ID", "Status"]
    )

    assert result is True
    assert confluence.last.call["filename"] == "orders-29-06-2026.xlsx"
    assert confluence.last.call["comment"] == "Total orders 2"
    assert confluence.last.call["page_id"] == "page-1"


async def test_async_infers_headers_and_rows(confluence):
    client = FakeClient(FakeAsyncOrdersService([{"id": "ORD-1", "status": "Querying"}]))

    result = await publish_xlsx_to_confluence_async(client, FakeSettings())

    assert result is True
    worksheet = load_workbook(BytesIO(confluence.last.call["file_content"])).active
    assert [cell.value for cell in worksheet[1]] == ["id", "status"]


async def test_async_rejects_mismatched_layout(confluence):
    client = FakeClient(FakeAsyncOrdersService([]))

    with pytest.raises(ValueError):
        await publish_xlsx_to_confluence_async(client, FakeSettings(), row_mapper=_to_row)
