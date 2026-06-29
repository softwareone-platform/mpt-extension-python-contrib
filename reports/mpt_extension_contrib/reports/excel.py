"""Build Excel workbooks for vendor reports."""

from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

type SheetDefinition = tuple[str, list[str], list[list[str]]]

EXCEL_MIME_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

_HEADER_FONT = Font(bold=True)


class ExcelReportBuilder:
    """Build an Excel workbook from report rows and return its raw bytes."""

    def __init__(self, headers: list[str], sheet_name: str = "Report") -> None:
        """Initialize the builder.

        Args:
            headers: Column headers written as the first, bold row.
            sheet_name: Title of the single sheet produced by ``build_sheet``.
        """
        self.headers = headers
        self.sheet_name = sheet_name

    def build_sheet(self, rows: list[list[str]]) -> bytes:
        """Build a single-sheet workbook in memory and return its raw bytes.

        Args:
            rows: Data rows appended below the header row.

        Returns:
            The workbook serialized to bytes.
        """
        workbook = Workbook()
        worksheet = workbook.active
        if worksheet.title == "Sheet":
            worksheet.title = self.sheet_name
        self._populate_sheet(worksheet, self.headers, rows)
        return self._save_to_bytes(workbook)

    def build_multi_sheet(self, sheets: list[SheetDefinition]) -> bytes:
        """Build a workbook with multiple named sheets and return its raw bytes.

        Args:
            sheets: One ``(title, headers, rows)`` tuple per sheet.

        Returns:
            The workbook serialized to bytes.

        Raises:
            ValueError: If ``sheets`` is empty.
        """
        if not sheets:
            raise ValueError("sheets must contain at least one sheet definition")
        workbook = Workbook()
        for idx, sheet_def in enumerate(sheets):
            worksheet = workbook.active if idx == 0 else workbook.create_sheet()
            worksheet.title = sheet_def[0]
            self._populate_sheet(worksheet, sheet_def[1], sheet_def[2])
        return self._save_to_bytes(workbook)

    def save(self, file_path: str, file_content: bytes) -> None:
        """Write workbook bytes to a file path.

        Args:
            file_path: Destination path.
            file_content: Workbook bytes to write.
        """
        Path(file_path).write_bytes(file_content)

    def _populate_sheet(
        self,
        worksheet: Worksheet,
        headers: list[str],
        rows: list[list[str]],
    ) -> None:
        worksheet.append(headers)
        for cell in worksheet[1]:
            cell.font = _HEADER_FONT
        for row_data in rows:
            worksheet.append(row_data)
        worksheet.auto_filter.ref = worksheet.dimensions

    def _save_to_bytes(self, workbook: Workbook) -> bytes:
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()
